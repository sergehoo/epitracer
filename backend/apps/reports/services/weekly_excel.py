"""Rendu du rapport hebdomadaire au format tableur (Excel ou CSV fallback).

Sortie primaire : XLSX via openpyxl (multi-feuilles avec styles).
Fallback : CSV plat via stdlib si openpyxl absent (dev env sans lib).

Feuilles du XLSX :
  1. Résumé          — période, KPI globaux
  2. Risques         — 4 niveaux × (nombre, %, évolution)
  3. Cas             — classification épidémiologique
  4. Laboratoire     — prélèvements + analyses
  5. Points d'entrée — top N
  6. Districts       — top N
  7. Maladies        — top N
  8. Comparaison     — vs semaine précédente

Sécurité — AC-06 : les valeurs commençant par =, +, -, @ sont préfixées
d'une apostrophe pour neutraliser les formules injectées (CSV injection).
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger("epidemitracker.reports.excel")

# Caractères qui déclenchent l'interprétation d'une formule dans Excel
CSV_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _neutralize(value: Any) -> Any:
    """AC-06 : neutralise les tentatives d'injection de formule Excel.

    Un attaquant qui saisirait `=cmd|'/c calc'!A0` dans un champ voyageur
    verrait sa formule exécutée quand un opérateur ouvre le fichier.
    On préfixe d'une apostrophe (marqueur "texte forcé" Excel).
    """
    if not isinstance(value, str):
        return value
    if value and value[0] in CSV_INJECTION_PREFIXES:
        return f"'{value}"
    return value


def render_weekly_xlsx(agg: dict) -> bytes:
    """Rend le rapport en XLSX si openpyxl dispo, sinon CSV fallback.

    Retourne toujours des bytes prêts à écrire dans un FileField.
    """
    try:
        return _render_xlsx_openpyxl(agg)
    except ImportError:
        logger.warning("openpyxl absent — fallback CSV plat.")
        return _render_csv_fallback(agg)


# ---------------------------------------------------------------------------
# XLSX primary — openpyxl (styles, plusieurs feuilles)
# ---------------------------------------------------------------------------
def _render_xlsx_openpyxl(agg: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    # Palette CI
    HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    ACCENT_FILL = PatternFill(start_color="F77F00", end_color="F77F00", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14, color="003366")

    def _style_header(ws, row_idx: int, n_cols: int):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _autosize(ws, max_width: int = 40):
        for col in ws.columns:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(length + 2, max_width)

    wb = Workbook()

    period = agg.get("period", {})
    meta = agg.get("meta", {})

    # ─── Feuille 1 : Résumé ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1["A1"] = "Rapport hebdomadaire EpiTrace"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:B1")
    ws1["A2"] = "Période"
    ws1["B2"] = _neutralize(period.get("label", ""))
    ws1["A3"] = "Semaine ISO"
    ws1["B3"] = f"{period.get('iso_year', '')} — S{period.get('iso_week', '')}"
    ws1["A4"] = "Généré le"
    ws1["B4"] = _neutralize(meta.get("generated_at", ""))[:19]
    ws1["A5"] = "Fuseau"
    ws1["B5"] = _neutralize(meta.get("tz", "Africa/Abidjan"))

    ws1.append([])
    ws1.append(["Indicateur", "Valeur"])
    _style_header(ws1, ws1.max_row, 2)
    for label, value in [
        ("Voyageurs enregistrés", agg.get("travelers", {}).get("registered", 0)),
        ("En suivi actif", agg.get("travelers", {}).get("active_followup", 0)),
        ("Nouveaux suivis", agg.get("followups", {}).get("new", 0)),
        ("Suivis terminés", agg.get("followups", {}).get("completed", 0)),
        ("Check-ins reçus", agg.get("checkins", {}).get("received", 0)),
        ("Check-ins manqués", agg.get("checkins", {}).get("missed", 0)),
        ("Demandes d'assistance", agg.get("assistance", {}).get("requests", 0)),
        ("Alertes créées", agg.get("alerts", {}).get("created", 0)),
        ("Alertes ouvertes", agg.get("alerts", {}).get("open", 0)),
        ("Alertes résolues", agg.get("alerts", {}).get("resolved", 0)),
    ]:
        ws1.append([_neutralize(label), value])
    _autosize(ws1)

    # ─── Feuille 2 : Risques ───────────────────────────────────────────
    ws2 = wb.create_sheet("Risques")
    ws2.append(["Niveau", "Nombre", "Pourcentage"])
    _style_header(ws2, 1, 3)
    risks = agg.get("risk_levels", {})
    for level, label in [
        ("critical", "Critique"),
        ("high", "Élevé"),
        ("moderate", "Modéré"),
        ("normal", "Normal"),
    ]:
        data = risks.get(level, {}) or {}
        ws2.append([label, data.get("count", 0), data.get("pct", 0)])
    ws2.append(["TOTAL", risks.get("total", 0), 100])
    _autosize(ws2)

    # ─── Feuille 3 : Cas ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Cas")
    ws3.append(["Classification", "Nombre"])
    _style_header(ws3, 1, 2)
    cases = agg.get("cases", {})
    for code, label in [
        ("suspect", "Suspects"), ("probable", "Probables"),
        ("confirmed", "Confirmés"), ("discarded", "Exclus"),
    ]:
        ws3.append([label, cases.get(code, 0)])
    _autosize(ws3)

    # ─── Feuille 4 : Laboratoire ───────────────────────────────────────
    ws4 = wb.create_sheet("Laboratoire")
    ws4.append(["Indicateur", "Valeur"])
    _style_header(ws4, 1, 2)
    samples = agg.get("samples", {})
    analyses = agg.get("analyses", {})
    for label, value in [
        ("Prélèvements demandés", samples.get("requested", 0)),
        ("Prélèvements réalisés", samples.get("collected", 0)),
        ("Analyses en attente", analyses.get("pending", 0)),
        ("Analyses en cours", analyses.get("in_progress", 0)),
        ("Analyses positives", analyses.get("positive", 0)),
        ("Analyses négatives", analyses.get("negative", 0)),
    ]:
        ws4.append([_neutralize(label), value])
    _autosize(ws4)

    # ─── Feuilles 5-7 : Breakdowns ─────────────────────────────────────
    for key, title in [
        ("by_entry_point", "Points d'entrée"),
        ("by_district", "Districts"),
        ("by_disease", "Maladies"),
    ]:
        ws = wb.create_sheet(title)
        ws.append(["Nom", "Nombre"])
        _style_header(ws, 1, 2)
        for row in agg.get(key, [])[:50]:
            ws.append([_neutralize(row.get("name", "—")), row.get("count", 0)])
        _autosize(ws)

    # ─── Feuille 8 : Comparaison ───────────────────────────────────────
    ws8 = wb.create_sheet("Comparaison")
    ws8.append(["Indicateur", "Actuel", "Précédent", "Évolution %"])
    _style_header(ws8, 1, 4)
    comp = agg.get("comparison", {})
    for key, label in [
        ("travelers", "Voyageurs enregistrés"),
        ("followups_new", "Nouveaux suivis"),
        ("checkins_received", "Check-ins reçus"),
        ("alerts_created", "Alertes créées"),
    ]:
        data = comp.get(key, {}) or {}
        ws8.append([
            _neutralize(label), data.get("current", 0),
            data.get("previous", 0), data.get("delta_pct", 0),
        ])
    _autosize(ws8)

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# CSV fallback — stdlib
# ---------------------------------------------------------------------------
def _render_csv_fallback(agg: dict) -> bytes:
    """Format CSV plat avec sections séparées par une ligne blanche.

    Utilise le séparateur point-virgule (norme FR — compat Excel FR sans
    changement de locale) + BOM UTF-8 pour l'ouverture native Excel.
    """
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)

    period = agg.get("period", {})
    writer.writerow(["Rapport hebdomadaire EpiTrace"])
    writer.writerow(["Période", _neutralize(period.get("label", ""))])
    writer.writerow([])

    writer.writerow(["=== KPI globaux ==="])
    writer.writerow(["Indicateur", "Valeur"])
    for label, value in [
        ("Voyageurs enregistrés", agg.get("travelers", {}).get("registered", 0)),
        ("En suivi actif", agg.get("travelers", {}).get("active_followup", 0)),
        ("Nouveaux suivis", agg.get("followups", {}).get("new", 0)),
        ("Check-ins reçus", agg.get("checkins", {}).get("received", 0)),
        ("Check-ins manqués", agg.get("checkins", {}).get("missed", 0)),
        ("Alertes ouvertes", agg.get("alerts", {}).get("open", 0)),
    ]:
        writer.writerow([_neutralize(label), value])
    writer.writerow([])

    writer.writerow(["=== Risques ==="])
    writer.writerow(["Niveau", "Nombre", "Pourcentage"])
    for level in ("critical", "high", "moderate", "normal"):
        d = agg.get("risk_levels", {}).get(level, {}) or {}
        writer.writerow([level, d.get("count", 0), d.get("pct", 0)])
    writer.writerow([])

    for key, title in [
        ("by_entry_point", "Points d'entrée"),
        ("by_district", "Districts"),
        ("by_disease", "Maladies"),
    ]:
        writer.writerow([f"=== {title} ==="])
        writer.writerow(["Nom", "Nombre"])
        for row in agg.get(key, [])[:50]:
            writer.writerow([_neutralize(row.get("name", "—")), row.get("count", 0)])
        writer.writerow([])

    # BOM UTF-8 pour Excel FR
    return ("﻿" + buf.getvalue()).encode("utf-8")
